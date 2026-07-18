# Génère public/downloads/template-okr-cosmo.xlsx — le template OKR gratuit
# téléchargeable (linkable asset, article /blog/template-okr-gratuit).
# Usage : python scripts/generate-okr-template.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SLATE = "0F172A"
BLUE = "2563EB"
LIGHT = "F1F5F9"
BORDER = Side(style="thin", color="CBD5E1")
THIN = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)

wb = Workbook()

# ── Feuille 1 : Mode d'emploi ────────────────────────────────────────────
ws = wb.active
ws.title = "Mode d'emploi"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 100

rows = [
    ("title", "Template OKR — par Cosmo (thecosmo.app)"),
    ("space", ""),
    ("body", "Ce template vous permet de définir et suivre vos OKR (Objectives & Key Results) sur un cycle de 12 semaines (un trimestre)."),
    ("space", ""),
    ("h2", "Comment l'utiliser"),
    ("body", "1. Choisissez 2 à 4 objectifs maximum pour le cycle — pas plus, c'est la règle qui fait fonctionner la méthode."),
    ("body", "2. Pour chaque objectif, définissez 2 à 5 résultats clés MESURABLES (un nombre, pas une intention)."),
    ("body", "3. Renseignez la valeur de départ et la valeur cible de chaque résultat clé."),
    ("body", "4. Chaque semaine (10 minutes suffisent), mettez à jour la colonne « Valeur actuelle » : la progression se calcule automatiquement."),
    ("body", "5. En fin de cycle, un OKR autour de 70 % est une réussite. 100 % partout = vous visiez trop bas."),
    ("space", ""),
    ("h2", "Rappels"),
    ("body", "• L'objectif est qualitatif et motivant (sans chiffre). Les chiffres vivent dans les résultats clés."),
    ("body", "• Un résultat clé mesure un résultat, pas une activité (« décrocher 5 entretiens » plutôt qu'« envoyer 50 candidatures »)."),
    ("body", "• Le suivi hebdomadaire est ce qui fait la différence entre un OKR et une résolution de janvier."),
    ("space", ""),
    ("body", "Guide complet et 15 exemples : https://thecosmo.app/blog/methode-okr-exemples"),
    ("body", "Marre du tableur ? Cosmo suit vos OKR automatiquement (gratuit) : https://thecosmo.app"),
]
r = 2
for kind, text in rows:
    c = ws.cell(row=r, column=2, value=text)
    if kind == "title":
        c.font = Font(size=16, bold=True, color=SLATE)
    elif kind == "h2":
        c.font = Font(size=12, bold=True, color=BLUE)
    else:
        c.font = Font(size=11, color="334155")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# ── Feuilles OKR ─────────────────────────────────────────────────────────
HEADERS = ["Objectif", "Résultat clé", "Valeur départ", "Valeur cible", "Valeur actuelle", "Progression", "Notes / prochaine action"]
WIDTHS = [34, 44, 14, 13, 15, 13, 40]

def okr_sheet(name, example_rows):
    s = wb.create_sheet(name)
    s.sheet_view.showGridLines = False
    s.freeze_panes = "A2"
    for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        cell = s.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=SLATE)
        cell.alignment = Alignment(vertical="center", horizontal="center" if i >= 3 else "left")
        cell.border = THIN
        s.column_dimensions[get_column_letter(i)].width = w
    s.row_dimensions[1].height = 22

    r = 2
    for obj, krs in example_rows:
        start = r
        for kr in krs:
            s.cell(row=r, column=2, value=kr[0])
            s.cell(row=r, column=3, value=kr[1])
            s.cell(row=r, column=4, value=kr[2])
            s.cell(row=r, column=5, value=kr[3])
            # Progression = (actuel - départ) / (cible - départ), bornée 0-100 %
            s.cell(row=r, column=6, value=f"=IF(D{r}=C{r},\"\",MAX(0,MIN(1,(E{r}-C{r})/(D{r}-C{r}))))")
            s.cell(row=r, column=6).number_format = "0%"
            for col in range(1, 8):
                cc = s.cell(row=r, column=col)
                cc.border = THIN
                cc.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center" if 3 <= col <= 6 else "left")
                if (r % 2) == 0:
                    cc.fill = PatternFill("solid", fgColor=LIGHT)
            r += 1
        s.merge_cells(start_row=start, start_column=1, end_row=r - 1, end_column=1)
        oc = s.cell(row=start, column=1, value=obj)
        oc.font = Font(bold=True, color=SLATE)
        oc.alignment = Alignment(wrap_text=True, vertical="top")
    # 2 objectifs vierges (4 KR chacun)
    for _ in range(2):
        start = r
        for _ in range(4):
            s.cell(row=r, column=6, value=f"=IF(D{r}=C{r},\"\",MAX(0,MIN(1,(E{r}-C{r})/(D{r}-C{r}))))")
            s.cell(row=r, column=6).number_format = "0%"
            for col in range(1, 8):
                cc = s.cell(row=r, column=col)
                cc.border = THIN
                cc.alignment = Alignment(wrap_text=True, vertical="center")
                if (r % 2) == 0:
                    cc.fill = PatternFill("solid", fgColor=LIGHT)
            r += 1
        s.merge_cells(start_row=start, start_column=1, end_row=r - 1, end_column=1)

okr_sheet("OKR personnels", [
    ("Retrouver une forme physique dont je suis fier", [
        ("Courir 3 séances par semaine pendant 12 semaines (36 séances)", 0, 36, 9),
        ("Passer de 0 à 10 km en continu", 0, 10, 4),
        ("Dormir 7 h 30 en moyenne par nuit (heures)", 6.5, 7.5, 7.0),
    ]),
])

okr_sheet("OKR équipe-startup", [
    ("Faire du contenu notre premier canal d'acquisition", [
        ("Visites organiques mensuelles", 2000, 8000, 3100),
        ("Articles SEO publiés", 0, 12, 3),
        ("Backlinks de sites d'autorité obtenus", 0, 15, 2),
    ]),
])

import os
os.makedirs("public/downloads", exist_ok=True)
out = "public/downloads/template-okr-cosmo.xlsx"
wb.save(out)
print("OK", out, os.path.getsize(out), "octets")
